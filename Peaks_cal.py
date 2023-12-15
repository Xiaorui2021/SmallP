"""
object：
Standardizing the output of Peaks by incorporating quality deviations, scan information, etc.
1 Parameter alignment
2 Standardizing PTM modifications using the Unimod database
3 Computing Cys framework information
4 Supplementing delta_M and Scan information
"""

import pandas as pd
import openpyxl
from openpyxl import Workbook
import os
import re
import time

PTM1="(-2.02)"
PTM2="(+57.02)"
PTM_2S="Disulphide bond formation"
PTM_Iaa="Carbamidomethylation"
seq_screen="C"
species="Pheretima aspergillum"
ppm_l=-5
ppm_r=10

def deta_M_match(file_path, file_path1,document1, type):
    workbook1 = openpyxl.load_workbook(file_path1 + '/'+"DL-TQY-fraction_"+ f'{type}.xlsx')
    worksheet1 = workbook1.get_sheet_by_name('Sheet1')
    workbook2 = openpyxl.load_workbook(file_path + '/'+ document1 + "_2C peptides_temp.xlsx")
    worksheet2 = workbook2.get_sheet_by_name('Sheet1')
    deta_m = 0.015
    deta_RT = 0.2

    wb1 = Workbook()
    ws1 = wb1.active
    max_row1 = worksheet1.max_row
    max_col1 = worksheet1.max_column
    print('max_row1:%s' % max_row1)
    print('max_col1:%s' % max_col1)

    data_ws1 = {}
    for row1 in worksheet1.iter_rows(min_row=2, max_col=max_col1, max_row=max_row1, values_only=True):
        file_1 = row1[max_col1 - 1]
        if file_1 not in data_ws1:
            data_ws1[file_1] = []
        temp = {}
        temp['RT_1'] = row1[6]
        temp['pepmass_1'] = row1[3]
        temp['deta_M_1'] = row1[5]
        temp['scan_1'] = row1[1]
        data_ws1[file_1].append(temp)

    wb2 = Workbook()
    ws2 = wb2.active
    max_row2 = worksheet2.max_row
    max_col2 = worksheet2.max_column
    print('max_row2:%s' % max_row2)
    print('max_col2:%s' % max_col2)

    for i in range(max_row2):
        rowi = [row.value for row in list(worksheet2.rows)[i]]
        for j in range(max_col2):
            aij = rowi[j]
            ws2.cell(row=i + 1, column=j + 1, value=aij)

    RT_2 = [i.value for i in list(ws2.columns)[26]]
    pepmass_2 = [i.value for i in list(ws2.columns)[24]]
    file_2_0 = [i.value for i in list(ws2.columns)[28]]

    for i in range(1, max_row2):
        file_2 = file_2_0[i].replace("_rev","").replace(".raw","")
        if file_2 in data_ws1:
            for j in range(0, len(data_ws1[file_2])):
                if abs(float(RT_2[i]) - float(data_ws1[file_2][j]['RT_1'])) < deta_RT and abs(
                        float(pepmass_2[i]) - float(data_ws1[file_2][j]['pepmass_1'])) < deta_m:
                    ws2.cell(row=i + 1, column=36,
                             value=f'{data_ws1[file_2][j]["deta_M_1"]}')  # 循环写入数据，value后面输入需要写入的值
                    ws2.cell(row=i + 1, column=28,
                             value=f'{data_ws1[file_2][j]["scan_1"]}')  # 循环写入数据，value后面输入需要写入的值
    wb2.save(file_path + '/' + document1 + '_identification.xlsx')

start_time = time.time()
file_path0= r'C:\Users\bqxin\Desktop\SmallP\Peaks_cal'
os.mkdir(file_path0+"/cal")
file_path= file_path0+"/"+"cal"
file_path1= r'C:\Users\bqxin\Desktop\SmallP\raw data'

# step1：Data import
print("step1：Data import")
file_name3 = os.listdir(file_path0)  # 读取该目录下所有文件名
file_name3 = [i for i in file_name3 if '.csv' in i]  # 只保存名称中含有.xls的文件
df = pd.read_excel(file_path1 + "/" + "The PTM database of variety software.xlsx", sheet_name="Peaks")
Udata_ls=[]
Name_ls=[]
for i in  range(len(df["NO."])):
    Unimod_accession=df["Unimod_accession"][i]
    Name = df["Name"][i]
    st_1 = Name.find('(')
    if st_1!=-1:
        PTM = Name[:st_1 -1].replace(","," ")
    else:
        PTM=Name.replace(","," ")
    Udata_ls.append(df["Unimod_accession"][i])
    Name_ls.append(PTM)
mod_dict=dict(zip(Name_ls, Udata_ls))

for file in file_name3:
    print(file)
    db=(re.split(r'[-_]', file))[0]
    type= (re.split(r'[-_]', file))[1]
    software = (re.split(r'[-_]', file))[2]
    Num_=(re.split(r'[-_]', file))[3].replace(".csv","")
    document0 = db + "-" + type + "_" + software+"_" + Num_
    document1 = species+"-"+db + "-" + type + "_" + software

    df = pd.read_csv(file_path0 + '/' + document0 + ".csv")
    col_name = df.columns.tolist()
    # 待增加的列
    add_list = ["_NO.", "_Compound", "_Sequence", "_length", "_Score", "_Mass",
                "_Ppm", "_m/z", "_Charge", "_RT", "_Scan", "_File", "_Software",
                "_PTM-Unimod", "_Accession", "_Sample", "_Area", "_Type", "_DetaM",
                "_Cnum", "_Bondnum", "_2S_Val", "_C_pos", "_C_frame", "_Cf_len", "_CF_Freq"]
    for i in range(len(add_list)):
        index = len(col_name) + i + 1
        col_name.insert(index, add_list[i])
        df = df.reindex(columns=col_name)
    df.to_excel(file_path + '/' + document1 + "_temp.xlsx", index=False)

    print("Step2: Parameter alignment")
    # Step2: Parameter alignment
    df = pd.read_excel(file_path + '/' + document1 + "_temp.xlsx", sheet_name=0)
    n = 0
    for i in range(len(df["Peptide"])):
        n += 1
        df["_NO."][i] = n
        df["_Compound"][i] = str(round(df["RT"][i], 2)) + "_" + str(round(df["Mass"][i], 2))
        df["_Sequence"][i] = re.sub(r"\(.*?\)", "", df["Peptide"][i])
        df["_length"][i] = len(df["_Sequence"][i])
        df["_Score"][i] = df["-10lgP"][i]
        df["_Ppm"][i] = df["ppm"][i]
        df["_Mass"][i] = df["Mass"][i]
        df["_Charge"][i] = round(df["Mass"][i] / df["m/z"][i])
        df["_m/z"][i] = df["m/z"][i]
        df["_RT"][i] = df["RT"][i]
        df["_File"][i] = df["Source File"][i].replace(".mgf","")
        df["_Software"][i] = software
        df["_Type"][i]=type
        df["_Accession"][i] = df["Accession"][i]
        sf = df["Source File"][i]
        if "_" in sf:
            st1_ = sf.index("_")
            df["_Sample"][i] = sf[:st1_].replace(f"-{type}", "")
        else:
            st1_ = sf.index(".")
            df["_Sample"][i] = sf[:st1_].replace(f"-{type}", "")

        df["_Area"][i] = df["Area"+" "+f"{type}"][i]
        df["_Cnum"][i] = df["Peptide"][i].count(seq_screen)
        if type=="Nat":
            df["_Bondnum"][i] = df["Peptide"][i].count(PTM1)
            df["_2S_Val"][i] = df["_Cnum"][i] - 2 * df["_Bondnum"][i]
        else:
            df["_Bondnum"][i] = df["Peptide"][i].count(PTM2)
            df["_2S_Val"][i] = df["_Cnum"][i] - df["_Bondnum"][i]

        mod_ls1 = []
        PTM = str(df["PTM"][i])
        a = PTM.split("; ")
        for j in range(len(a)):
            st_1 = a[j].find('(')
            if st_1 != -1:
                mod = a[j][:st_1 - 1]
            else:
                mod = a[j]
            if type == "Nat":
                if mod != f"{PTM_2S}" and mod != "nan":
                    Peaks_unimod = mod_dict[str(mod)]
                    mod_ls1.append(Peaks_unimod)
            else:
                if mod != f"{PTM_Iaa}" and mod != "nan":
                    Peaks_unimod = mod_dict[str(mod)]
                    mod_ls1.append(Peaks_unimod)
        df["_PTM-Unimod"][i] = str(mod_ls1)

        st_1 = df["_Sequence"][i].find(seq_screen)
        a = []
        b = []
        c = 0
        if st_1 != -1 and st_1 != len(df["_Sequence"][i]) - len(seq_screen):
            a = [m.start() for m in re.finditer(seq_screen, df["_Sequence"][i])]
            b = []
            j = 0
            while j < len(a) - 1:
                c = a[j + 1] - a[j]
                b.append(c)
                j += 1
        df["_C_pos"][i] = str(a)
        df["_C_frame"][i] = str(b)
        df["_Cf_len"][i] = str(len(b))
    df.to_excel(file_path + '/' + document1 + "_identification_rev_temp.xlsx", index=False)
    df4 = pd.read_excel(file_path + '/' + document1 + "_identification_rev_temp.xlsx")
    if type=="Nat":
        df5 = df4[(df4["_2S_Val"] < 2) & (-1 < df4["_2S_Val"]) & (1 < df4["_Cnum"]) & (df4["_Ppm"] < ppm_r) & (ppm_l < df4["_Ppm"])]
    else:
        df5 = df4[(df4["_Ppm"] < ppm_r) & (ppm_l < df4["_Ppm"])]
    df5.to_excel(file_path + "/" + document1 + "_2C peptides_temp.xlsx", index=False)

    print("Step 3: Supplementing delta_M and Scan information")
    # Step 3: Supplementing delta_M and Scan information
    deta_M_match(file_path,file_path1, document1, type)

# step4: Delete excess files
print("Step4: Delete excess files")
fname = os.listdir(file_path)
fname = [i for i in fname if '_temp' in i]
for i in fname:
    os.remove(file_path + "/" + i)

# Computational time consumption
end_time = time.time()
escape = (end_time - start_time)/60
print('The code run {:.0f}m {:.0f}s'.format(escape // 60, escape % 60))
print("程序结束")
#————————————————
"""
Qirui Bi
Shanghai
December 7, 2023
"""